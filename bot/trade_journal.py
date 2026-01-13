"""
Trade Journal Export Module.

Provides functionality to export trade history to CSV and Excel formats
with comprehensive attribution data and analytics.
"""

from __future__ import annotations

import csv
import io
import json
import logging
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional, Union

logger = logging.getLogger(__name__)


@dataclass
class TradeEntry:
    """Single trade entry for the journal."""
    trade_id: str
    timestamp: datetime
    symbol: str
    side: Literal["BUY", "SELL"]
    entry_price: float
    exit_price: Optional[float] = None
    quantity: float = 0.0
    pnl: float = 0.0
    pnl_percent: float = 0.0
    fees: float = 0.0
    strategy: str = "unknown"
    model: str = "unknown"
    regime: str = "unknown"
    confidence: float = 0.0
    signal_strength: float = 0.0
    market_type: str = "crypto"
    holding_period_hours: float = 0.0
    notes: str = ""
    tags: List[str] = field(default_factory=list)
    exit_reason: str = ""
    slippage: float = 0.0
    max_drawdown_during: float = 0.0
    risk_reward_ratio: float = 0.0
    stop_loss: Optional[float] = None
    take_profit: Optional[float] = None


class TradeJournal:
    """
    Trade journal manager for exporting trade history.

    Supports CSV and Excel export with full attribution data,
    analytics, and performance summaries.
    """

    def __init__(
        self,
        db_path: str = "data/trading.db",
        export_dir: str = "data/exports",
    ):
        self.db_path = Path(db_path)
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)

        self._trades: List[TradeEntry] = []

    def load_trades_from_db(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        symbol: Optional[str] = None,
        strategy: Optional[str] = None,
    ) -> List[TradeEntry]:
        """
        Load trades from database with optional filters.

        Args:
            start_date: Filter trades after this date
            end_date: Filter trades before this date
            symbol: Filter by symbol
            strategy: Filter by strategy name

        Returns:
            List of TradeEntry objects
        """
        try:
            import sqlite3

            if not self.db_path.exists():
                logger.warning(f"Database not found: {self.db_path}")
                return []

            conn = sqlite3.connect(str(self.db_path))
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            # Build query
            query = "SELECT * FROM trades WHERE 1=1"
            params = []

            if start_date:
                query += " AND timestamp >= ?"
                params.append(start_date.isoformat())

            if end_date:
                query += " AND timestamp <= ?"
                params.append(end_date.isoformat())

            if symbol:
                query += " AND symbol = ?"
                params.append(symbol)

            if strategy:
                query += " AND strategy = ?"
                params.append(strategy)

            query += " ORDER BY timestamp DESC"

            cursor.execute(query, params)
            rows = cursor.fetchall()

            trades = []
            for row in rows:
                trade = TradeEntry(
                    trade_id=str(row.get("id", "")),
                    timestamp=datetime.fromisoformat(row["timestamp"]) if row.get("timestamp") else datetime.now(),
                    symbol=row.get("symbol", ""),
                    side=row.get("side", "BUY"),
                    entry_price=float(row.get("entry_price", 0)),
                    exit_price=float(row["exit_price"]) if row.get("exit_price") else None,
                    quantity=float(row.get("quantity", 0)),
                    pnl=float(row.get("pnl", 0)),
                    pnl_percent=float(row.get("pnl_percent", 0)),
                    fees=float(row.get("fees", 0)),
                    strategy=row.get("strategy", "unknown"),
                    model=row.get("model", "unknown"),
                    regime=row.get("regime", "unknown"),
                    confidence=float(row.get("confidence", 0)),
                    signal_strength=float(row.get("signal_strength", 0)),
                    market_type=row.get("market_type", "crypto"),
                    notes=row.get("notes", ""),
                    exit_reason=row.get("exit_reason", ""),
                )
                trades.append(trade)

            conn.close()
            self._trades = trades
            return trades

        except Exception as e:
            logger.error(f"Error loading trades from database: {e}")
            return []

    def load_trades_from_json(self, json_path: str) -> List[TradeEntry]:
        """Load trades from a JSON file."""
        try:
            path = Path(json_path)
            if not path.exists():
                logger.warning(f"JSON file not found: {path}")
                return []

            with open(path) as f:
                data = json.load(f)

            trades = []
            trade_list = data if isinstance(data, list) else data.get("trades", [])

            for item in trade_list:
                trade = TradeEntry(
                    trade_id=str(item.get("id", item.get("trade_id", ""))),
                    timestamp=datetime.fromisoformat(item["timestamp"]) if item.get("timestamp") else datetime.now(),
                    symbol=item.get("symbol", ""),
                    side=item.get("side", "BUY"),
                    entry_price=float(item.get("entry_price", 0)),
                    exit_price=float(item["exit_price"]) if item.get("exit_price") else None,
                    quantity=float(item.get("quantity", item.get("amount", 0))),
                    pnl=float(item.get("pnl", 0)),
                    pnl_percent=float(item.get("pnl_percent", 0)),
                    fees=float(item.get("fees", 0)),
                    strategy=item.get("strategy", "unknown"),
                    model=item.get("model", "unknown"),
                    regime=item.get("regime", "unknown"),
                    confidence=float(item.get("confidence", 0)),
                    signal_strength=float(item.get("signal_strength", 0)),
                    market_type=item.get("market_type", "crypto"),
                    notes=item.get("notes", ""),
                    exit_reason=item.get("exit_reason", ""),
                    tags=item.get("tags", []),
                )
                trades.append(trade)

            self._trades = trades
            return trades

        except Exception as e:
            logger.error(f"Error loading trades from JSON: {e}")
            return []

    def add_trade(self, trade: TradeEntry) -> None:
        """Add a trade to the journal."""
        self._trades.append(trade)

    def export_to_csv(
        self,
        filename: Optional[str] = None,
        trades: Optional[List[TradeEntry]] = None,
        include_summary: bool = True,
    ) -> str:
        """
        Export trades to CSV format.

        Args:
            filename: Output filename (auto-generated if not provided)
            trades: Trades to export (uses loaded trades if not provided)
            include_summary: Include summary statistics at the bottom

        Returns:
            Path to exported file
        """
        trades = trades or self._trades

        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"trade_journal_{timestamp}.csv"

        filepath = self.export_dir / filename

        # Define columns
        columns = [
            "Trade ID", "Timestamp", "Symbol", "Side", "Market Type",
            "Entry Price", "Exit Price", "Quantity", "PnL ($)", "PnL (%)",
            "Fees", "Strategy", "Model", "Regime", "Confidence",
            "Signal Strength", "Holding Period (hrs)", "Exit Reason",
            "Stop Loss", "Take Profit", "Risk/Reward", "Slippage",
            "Max Drawdown During", "Notes", "Tags"
        ]

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            # Header
            writer.writerow(columns)

            # Trades
            for trade in trades:
                writer.writerow([
                    trade.trade_id,
                    trade.timestamp.isoformat() if trade.timestamp else "",
                    trade.symbol,
                    trade.side,
                    trade.market_type,
                    f"{trade.entry_price:.8f}",
                    f"{trade.exit_price:.8f}" if trade.exit_price else "",
                    f"{trade.quantity:.8f}",
                    f"{trade.pnl:.2f}",
                    f"{trade.pnl_percent:.2f}",
                    f"{trade.fees:.4f}",
                    trade.strategy,
                    trade.model,
                    trade.regime,
                    f"{trade.confidence:.2f}",
                    f"{trade.signal_strength:.2f}",
                    f"{trade.holding_period_hours:.1f}",
                    trade.exit_reason,
                    f"{trade.stop_loss:.8f}" if trade.stop_loss else "",
                    f"{trade.take_profit:.8f}" if trade.take_profit else "",
                    f"{trade.risk_reward_ratio:.2f}",
                    f"{trade.slippage:.4f}",
                    f"{trade.max_drawdown_during:.2f}",
                    trade.notes,
                    ";".join(trade.tags) if trade.tags else "",
                ])

            if include_summary and trades:
                # Add summary section
                writer.writerow([])
                writer.writerow(["=" * 20, "SUMMARY", "=" * 20])
                writer.writerow([])

                summary = self.calculate_summary(trades)
                for key, value in summary.items():
                    writer.writerow([key, value])

        logger.info(f"Exported {len(trades)} trades to {filepath}")
        return str(filepath)

    def export_to_csv_string(self, trades: Optional[List[TradeEntry]] = None) -> str:
        """Export trades to CSV string (for API response)."""
        trades = trades or self._trades
        output = io.StringIO()

        columns = [
            "Trade ID", "Timestamp", "Symbol", "Side", "Market Type",
            "Entry Price", "Exit Price", "Quantity", "PnL ($)", "PnL (%)",
            "Fees", "Strategy", "Model", "Regime", "Confidence",
            "Signal Strength", "Exit Reason", "Notes"
        ]

        writer = csv.writer(output)
        writer.writerow(columns)

        for trade in trades:
            writer.writerow([
                trade.trade_id,
                trade.timestamp.isoformat() if trade.timestamp else "",
                trade.symbol,
                trade.side,
                trade.market_type,
                f"{trade.entry_price:.8f}",
                f"{trade.exit_price:.8f}" if trade.exit_price else "",
                f"{trade.quantity:.8f}",
                f"{trade.pnl:.2f}",
                f"{trade.pnl_percent:.2f}",
                f"{trade.fees:.4f}",
                trade.strategy,
                trade.model,
                trade.regime,
                f"{trade.confidence:.2f}",
                f"{trade.signal_strength:.2f}",
                trade.exit_reason,
                trade.notes,
            ])

        return output.getvalue()

    def export_to_excel(
        self,
        filename: Optional[str] = None,
        trades: Optional[List[TradeEntry]] = None,
        include_charts: bool = True,
    ) -> str:
        """
        Export trades to Excel format with multiple sheets.

        Args:
            filename: Output filename
            trades: Trades to export
            include_charts: Include chart sheets (requires openpyxl)

        Returns:
            Path to exported file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV export")
            return self.export_to_csv(filename, trades)

        trades = trades or self._trades

        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"trade_journal_{timestamp}.xlsx"

        filepath = self.export_dir / filename

        wb = openpyxl.Workbook()

        # Sheet 1: Trades
        ws_trades = wb.active
        ws_trades.title = "Trades"

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        profit_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        loss_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Headers
        headers = [
            "Trade ID", "Timestamp", "Symbol", "Side", "Market Type",
            "Entry Price", "Exit Price", "Quantity", "PnL ($)", "PnL (%)",
            "Fees", "Strategy", "Model", "Regime", "Confidence",
            "Signal Strength", "Exit Reason"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws_trades.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows
        for row_idx, trade in enumerate(trades, 2):
            data = [
                trade.trade_id,
                trade.timestamp.strftime("%Y-%m-%d %H:%M") if trade.timestamp else "",
                trade.symbol,
                trade.side,
                trade.market_type,
                trade.entry_price,
                trade.exit_price or "",
                trade.quantity,
                trade.pnl,
                trade.pnl_percent,
                trade.fees,
                trade.strategy,
                trade.model,
                trade.regime,
                trade.confidence,
                trade.signal_strength,
                trade.exit_reason,
            ]

            for col, value in enumerate(data, 1):
                cell = ws_trades.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

                # Color PnL cells
                if col == 9:  # PnL ($) column
                    if isinstance(value, (int, float)):
                        if value > 0:
                            cell.fill = profit_fill
                        elif value < 0:
                            cell.fill = loss_fill

        # Auto-width columns
        for col in range(1, len(headers) + 1):
            ws_trades.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

        # Sheet 2: Summary
        ws_summary = wb.create_sheet("Summary")
        summary = self.calculate_summary(trades)

        ws_summary.cell(row=1, column=1, value="Trade Journal Summary").font = Font(bold=True, size=14)

        row = 3
        for key, value in summary.items():
            ws_summary.cell(row=row, column=1, value=key)
            ws_summary.cell(row=row, column=2, value=value)
            row += 1

        # Sheet 3: By Strategy
        ws_strategy = wb.create_sheet("By Strategy")
        strategy_stats = self.calculate_by_strategy(trades)

        ws_strategy.cell(row=1, column=1, value="Performance by Strategy").font = Font(bold=True, size=14)

        strategy_headers = ["Strategy", "Trades", "Win Rate", "Total PnL", "Avg PnL", "Profit Factor"]
        for col, header in enumerate(strategy_headers, 1):
            cell = ws_strategy.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row = 4
        for strategy, stats in strategy_stats.items():
            ws_strategy.cell(row=row, column=1, value=strategy)
            ws_strategy.cell(row=row, column=2, value=stats["trades"])
            ws_strategy.cell(row=row, column=3, value=f"{stats['win_rate']:.1f}%")
            ws_strategy.cell(row=row, column=4, value=f"${stats['total_pnl']:.2f}")
            ws_strategy.cell(row=row, column=5, value=f"${stats['avg_pnl']:.2f}")
            ws_strategy.cell(row=row, column=6, value=f"{stats['profit_factor']:.2f}")
            row += 1

        # Sheet 4: Daily Performance
        ws_daily = wb.create_sheet("Daily Performance")
        daily_stats = self.calculate_daily_performance(trades)

        ws_daily.cell(row=1, column=1, value="Daily Performance").font = Font(bold=True, size=14)

        daily_headers = ["Date", "Trades", "PnL ($)", "Win Rate", "Best Trade", "Worst Trade"]
        for col, header in enumerate(daily_headers, 1):
            cell = ws_daily.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row = 4
        for date, stats in daily_stats.items():
            ws_daily.cell(row=row, column=1, value=date)
            ws_daily.cell(row=row, column=2, value=stats["trades"])
            ws_daily.cell(row=row, column=3, value=f"${stats['pnl']:.2f}")
            ws_daily.cell(row=row, column=4, value=f"{stats['win_rate']:.1f}%")
            ws_daily.cell(row=row, column=5, value=f"${stats['best']:.2f}")
            ws_daily.cell(row=row, column=6, value=f"${stats['worst']:.2f}")
            row += 1

        wb.save(filepath)
        logger.info(f"Exported {len(trades)} trades to {filepath}")
        return str(filepath)

    def calculate_summary(self, trades: List[TradeEntry]) -> Dict[str, Any]:
        """Calculate summary statistics for trades."""
        if not trades:
            return {"Total Trades": 0}

        total_trades = len(trades)
        winning_trades = [t for t in trades if t.pnl > 0]
        losing_trades = [t for t in trades if t.pnl < 0]

        total_pnl = sum(t.pnl for t in trades)
        total_fees = sum(t.fees for t in trades)
        gross_profit = sum(t.pnl for t in winning_trades) if winning_trades else 0
        gross_loss = abs(sum(t.pnl for t in losing_trades)) if losing_trades else 0

        win_rate = (len(winning_trades) / total_trades * 100) if total_trades > 0 else 0
        avg_win = (gross_profit / len(winning_trades)) if winning_trades else 0
        avg_loss = (gross_loss / len(losing_trades)) if losing_trades else 0
        profit_factor = (gross_profit / gross_loss) if gross_loss > 0 else float("inf")

        # Date range
        dates = [t.timestamp for t in trades if t.timestamp]
        date_range = f"{min(dates).strftime('%Y-%m-%d')} to {max(dates).strftime('%Y-%m-%d')}" if dates else "N/A"

        # Average holding period
        holding_periods = [t.holding_period_hours for t in trades if t.holding_period_hours > 0]
        avg_holding = sum(holding_periods) / len(holding_periods) if holding_periods else 0

        return {
            "Date Range": date_range,
            "Total Trades": total_trades,
            "Winning Trades": len(winning_trades),
            "Losing Trades": len(losing_trades),
            "Win Rate": f"{win_rate:.1f}%",
            "Total PnL": f"${total_pnl:.2f}",
            "Total Fees": f"${total_fees:.2f}",
            "Net PnL": f"${total_pnl - total_fees:.2f}",
            "Gross Profit": f"${gross_profit:.2f}",
            "Gross Loss": f"${gross_loss:.2f}",
            "Profit Factor": f"{profit_factor:.2f}" if profit_factor != float("inf") else "∞",
            "Average Win": f"${avg_win:.2f}",
            "Average Loss": f"${avg_loss:.2f}",
            "Average Holding Period": f"{avg_holding:.1f} hours",
            "Largest Win": f"${max(t.pnl for t in trades):.2f}",
            "Largest Loss": f"${min(t.pnl for t in trades):.2f}",
            "Average Trade PnL": f"${total_pnl / total_trades:.2f}",
        }

    def calculate_by_strategy(self, trades: List[TradeEntry]) -> Dict[str, Dict[str, Any]]:
        """Calculate performance metrics grouped by strategy."""
        strategy_trades: Dict[str, List[TradeEntry]] = {}

        for trade in trades:
            strategy = trade.strategy or "unknown"
            if strategy not in strategy_trades:
                strategy_trades[strategy] = []
            strategy_trades[strategy].append(trade)

        results = {}
        for strategy, strades in strategy_trades.items():
            wins = [t for t in strades if t.pnl > 0]
            losses = [t for t in strades if t.pnl < 0]
            total_pnl = sum(t.pnl for t in strades)
            gross_profit = sum(t.pnl for t in wins) if wins else 0
            gross_loss = abs(sum(t.pnl for t in losses)) if losses else 0

            results[strategy] = {
                "trades": len(strades),
                "win_rate": (len(wins) / len(strades) * 100) if strades else 0,
                "total_pnl": total_pnl,
                "avg_pnl": total_pnl / len(strades) if strades else 0,
                "profit_factor": gross_profit / gross_loss if gross_loss > 0 else float("inf"),
            }

        return results

    def calculate_daily_performance(self, trades: List[TradeEntry]) -> Dict[str, Dict[str, Any]]:
        """Calculate daily performance metrics."""
        daily_trades: Dict[str, List[TradeEntry]] = {}

        for trade in trades:
            if trade.timestamp:
                date = trade.timestamp.strftime("%Y-%m-%d")
                if date not in daily_trades:
                    daily_trades[date] = []
                daily_trades[date].append(trade)

        results = {}
        for date in sorted(daily_trades.keys()):
            dtrades = daily_trades[date]
            wins = [t for t in dtrades if t.pnl > 0]
            pnls = [t.pnl for t in dtrades]

            results[date] = {
                "trades": len(dtrades),
                "pnl": sum(pnls),
                "win_rate": (len(wins) / len(dtrades) * 100) if dtrades else 0,
                "best": max(pnls) if pnls else 0,
                "worst": min(pnls) if pnls else 0,
            }

        return results

    def get_export_formats(self) -> List[Dict[str, str]]:
        """Get available export formats."""
        return [
            {"format": "csv", "description": "Comma-separated values", "extension": ".csv"},
            {"format": "excel", "description": "Microsoft Excel", "extension": ".xlsx"},
            {"format": "json", "description": "JSON format", "extension": ".json"},
        ]

    def export_to_json(self, filename: Optional[str] = None, trades: Optional[List[TradeEntry]] = None) -> str:
        """Export trades to JSON format."""
        trades = trades or self._trades

        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"trade_journal_{timestamp}.json"

        filepath = self.export_dir / filename

        data = {
            "exported_at": datetime.now().isoformat(),
            "total_trades": len(trades),
            "summary": self.calculate_summary(trades),
            "trades": [
                {
                    "trade_id": t.trade_id,
                    "timestamp": t.timestamp.isoformat() if t.timestamp else None,
                    "symbol": t.symbol,
                    "side": t.side,
                    "market_type": t.market_type,
                    "entry_price": t.entry_price,
                    "exit_price": t.exit_price,
                    "quantity": t.quantity,
                    "pnl": t.pnl,
                    "pnl_percent": t.pnl_percent,
                    "fees": t.fees,
                    "strategy": t.strategy,
                    "model": t.model,
                    "regime": t.regime,
                    "confidence": t.confidence,
                    "signal_strength": t.signal_strength,
                    "holding_period_hours": t.holding_period_hours,
                    "exit_reason": t.exit_reason,
                    "notes": t.notes,
                    "tags": t.tags,
                }
                for t in trades
            ],
        }

        with open(filepath, "w") as f:
            json.dump(data, f, indent=2)

        logger.info(f"Exported {len(trades)} trades to {filepath}")
        return str(filepath)
